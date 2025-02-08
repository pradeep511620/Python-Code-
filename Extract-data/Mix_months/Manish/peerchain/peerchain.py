import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import pandas as pd
import openpyxl

sess = requests.Session()

base_url = 'https://www.peerchain.com/products/'


# product_urls = []
# product_url = open("peerchain_product_urls.csv", 'a', encoding='utf-8')
# product_data = open("peerchain_product_data.csv", 'a', encoding='utf-8')


def request_func(url):
    """Send Requests and Get Response from webpage and return Soup"""
    res = sess.get(url)
    soup = BeautifulSoup(res.content, 'lxml')
    return soup


def urljoin_func(url):
    """function takes a relative url and convert it into absolute url and returns"""
    return urljoin(base_url, url)


def extract_urls_recursive(url, visited=None):
    """
    Extracts all URLs from a webpage and any pages it links to, recursively.
    :param url: the URL of the starting page
    :param visited: a set of URLs that have already been visited
    :return: a set of all unique URLs found on the pages
    """
    if visited is None:
        visited = set()
    urls = set()
    try:
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            try:
                product_links = soup.find(class_="products columns-4").findAll('li')
            except AttributeError:
                product_links = soup.find(id="pg-w62a786bed04f5-1").findAll(class_='textwidget custom-html-widget')
            for link in product_links:
                next_url = urljoin_func(link.find('a')['href'])
                if next_url not in visited:
                    visited.add(next_url)
                    urls.add(next_url)
                    print(next_url)
                    urls = urls.union(extract_urls_recursive(next_url, visited))
    except Exception as e:
        print(e)
    return urls


def product_link(url):
    """find all the product links and save as csv file"""
    soup = request_func(url)
    product_links = soup.find(class_="enable-prdouct-variation-outer variation-table print-only").find('tr').findAll(class_='variation-data')
    for link in product_links:
        product_urls.append(urljoin_func(link.find('a')['href']))
        print(urljoin_func(link.find('a')['href']))
    # df = pd.DataFrame(product_urls)
    # df.to_csv(product_url, lineterminator='\n', index=False, header=False)


def product_details(url):
    """Get product details(Respective Product url, title, breadcrumb, grainer_sku, image url
    features, product details, product specifications) and call data save function(file as txt or csv)"""

    cols = [
        "S.No", "Mpn", "Label", "Upc", "Grainger_Sku", "Entity Id", "L3_Name", "L3_Entity_ID", "Item_Name", "item_ID",
        "Parent_name", "Parent_Name_Id", "Product_title", "Accessories", "Kits/Components", "Spare_Parts",
        "Product_Detail", "Uses", "Features", "Working_Mechanism", "Standards_and_Approvals", "Installation",
        "Accessories", "Image_Name", "Image_URL_1", "Image_URL_2", "Image_URL_3", "Image_URL_4", "Image_URL_5",
        "Video_Url", "Datasheet", "Shipping_length(mm)", "Shipping_height(mm)", "Shipping_width(mm)", "weight(kg)",
        "Quantity", "Price(usd)", "Discount(%)", "Country_of_Origin", "Cross_Reference", "Url", "Remarks", "attr_name",
        "attr_value"
    ]

    row = []

    soup = request_func(url)
    title = soup.h1.text.strip()
    print("title >>> ", title)
    category = soup.find(class_='product_meta').find(class_='posted_in').text
    print("Categories >>>>>>>>>>> ", category)
    try:
        grainer_sku = soup.find(class_='product_meta').find(class_='sku_wrapper').text
        print("sku >>>>>>>>>>> ", grainer_sku)
    except AttributeError:
        grainer_sku = ''
    try:
        description = soup.find(class_='top-description-product').text.replace('\n', ' ')
        print("description >>> ", description)
    except AttributeError:
        description = ''

    try:
        features = soup.find(class_='main-description-product').text.replace('\n', ' ')
        print("features >>> ", features)
    except AttributeError:
        features = ''

    image_list = [None] * 5
    try:
        images = soup.find(class_="woocommerce-product-gallery__wrapper print-only").findAll('a')
        for i, image in enumerate(images, start=0):
            image_list[i] = urljoin_func(image['href'])
        print("image >>>>>>>>>>>>", image_list)
    except Exception as e:
        print(e)

    attr_name = []
    attr_value = []
    try:
        table = soup.find('table', {"class": "print-only"}).findAll('tr')
        for td in table:
            attr_name.append(td.find(class_='variation-label').text)
            attr_value.append(f"rp_{td.find(class_='variation-data').text}")

        print("attr_name >>>>>>>>>>>>> ", attr_name)
        print("attr_value >>>>>>>>>>>>> ", attr_value)

    except Exception as e:
        print(e)

    """Convert data into dictionary form for pandas"""
    row.append({
        "Url": url,
        "Item_Name": title,
        "L3_Name": category,
        "Grainger_Sku": grainer_sku,
        "Product_Detail": description.strip(),
        "Features": features.strip(),
        "Image_URL_1": image_list[0],
        "Image_URL_2": image_list[1],
        "Image_URL_3": image_list[2],
        "Image_URL_4": image_list[3],
        "Image_URL_5": image_list[4],
        "attr_name": attr_name,
        "attr_value": attr_value
    })

    data_save(row, cols)


def data_save(row, cols):
    """function takes row and columns convert into dataframe and filter out and save data into the file"""

    df = pd.DataFrame(row, columns=cols)

    """Creating subset from Main Dataframe for specifications and drop specs"""
    specs_dataset = df[["Url", "Grainger_Sku", 'attr_name', 'attr_value']]
    specs_dataset = specs_dataset.explode(['attr_name', 'attr_value'])
    df = df.drop(['attr_name', 'attr_value'], axis=1)

    try:
        writer = pd.ExcelWriter('peerchain_product_data.xlsx', engine="openpyxl", mode='a', if_sheet_exists='overlay')
        source_file = openpyxl.load_workbook("peerchain_product_data.xlsx", enumerate)
        """Get number of rows in excel file (to determine where to append)"""
        product_sheet, specs_sheet = source_file["product_data_sheet"], source_file["product_specs_sheet"]
        row_count, row_count_spec = product_sheet.max_row, specs_sheet.max_row

        """writing data in Excel in two separated sheets(product_data_sheet, product_specs_sheet)"""
        df.to_excel(writer, sheet_name="product_data_sheet", index=False, startrow=row_count, header=False)
        specs_dataset.to_excel(writer, sheet_name="product_specs_sheet", index=False, startrow=row_count_spec, header=False)
        writer.close()
    except FileNotFoundError:
        writer = pd.ExcelWriter('peerchain_product_data.xlsx', engine="openpyxl", mode='w')
        df.to_excel(writer, sheet_name="product_data_sheet", index=False)
        specs_dataset.to_excel(writer, sheet_name="product_specs_sheet", index=False)
        writer.close()


def main():
    """Read urls from peerchain_product_urls.csv and calling product_urls one by one"""
    i = 2025
    product_urls = pd.read_csv('peerchain_product_urls.csv')['product_url']
    """using i indexing product url default i value is 0"""
    for url in product_urls[i:]:
        print(i, ">>>>>>>>>>>>>", url)
        product_details(url)
        i += 1


main()

"""Testing Call"""
# product_details('https://www.peerchain.com/product/100r-roller-chain/')

# pro_url = extract_urls_recursive(base_url)
# for url in pro_url:
#     try:
#         product_link(url)
#     except AttributeError as e:
#         print(e)


